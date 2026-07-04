#!/usr/bin/env python3
"""
Phase 1 UI Test File Inventory Orchestrator - v14

This script:
  1. Reads included repositories from a CSV/XLSX/JSONL file.
  2. Clones each repository locally.
  3. Calls the Node/ts-morph analyzer for each repo.
  4. Merges per-repo UI test file inventory results into CSV/JSONL outputs.

Default behavior:
  Ignores any existing `per_repo_outputs/{repo}.json`. The Node analyzer is
  always re-run and per-repo outputs are overwritten with fresh results. Clones
  in `--repo-cache` are still reused (driven by `clone_or_update_repo`).

With --resume:
  For each repo, look for a cached `per_repo_outputs/{repo}.json` and a sibling
  `per_repo_outputs/{repo}.meta.json`. If both exist, are well-formed, and the
  metadata matches the *current* analyzer config (analyzer file path + mtime
  and `--include-low-confidence` flag), skip clone + analyzer for that repo
  and aggregate the cached result. Otherwise re-run.

  Stale-cache cases (file changed, flags changed, corrupted JSON) are logged
  with a clear reason and counted in `overall_summary.json`.

With --refresh-clones:
  Invalidates both the cached clone and the cached per-repo JSON+meta. Forces
  a clean redo. Implies --no-resume in practice: even with --resume set,
  invalidated caches are not consulted because they get deleted first.

Per-repo summary rows include `analysis_source` ("cached" | "fresh") so
downstream analysis can see which repos were reused versus re-run.

This is Phase 1 only:
  - It outputs detected UI test files with evidence and confidence.
  - It also outputs setup/helper/support files, template files, and low-confidence candidates separately.
  - It does NOT extract RQ-level fixtures, inputs, actions, or assertions.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


INCLUDE_LABELS = {"web_application", "ui_component_library", "ui_widget_library"}


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("-", "_").replace(" ", "_")


def parse_full_name_from_url(url: str) -> str:
    marker = "github.com/"
    if marker not in url:
        return ""
    rest = url.split(marker, 1)[1].strip().strip("/")
    parts = rest.split("/")
    if len(parts) < 2:
        return ""
    owner, repo = parts[0], parts[1].replace(".git", "")
    return f"{owner}/{repo}"


def safe_repo_dir(full_name: str) -> str:
    return full_name.replace("/", "__").replace(":", "_")


def read_rows(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(path)

    if path.suffix.lower() in {".jsonl"}:
        rows = []
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rows.append(json.loads(line))
        return rows

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Reading XLSX requires openpyxl: pip install openpyxl") from exc

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
        return rows

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f)]


def get_repo_identity(row: Dict[str, Any]) -> Optional[Tuple[str, str]]:
    full_name = str(row.get("full_name") or "").strip()
    html_url = str(row.get("html_url") or "").strip()

    if not full_name and html_url:
        full_name = parse_full_name_from_url(html_url)

    if not html_url and full_name:
        html_url = f"https://github.com/{full_name}"

    if not full_name:
        return None
    return full_name, html_url


def should_include_repo(row: Dict[str, Any], only_included: bool) -> bool:
    if not only_included:
        return True

    for col in ("llm_include_exclude", "manual_include_exclude", "include_exclude"):
        v = normalize_label(row.get(col))
        if v == "include":
            return True
        if v == "exclude":
            return False

    for col in ("llm_project_subject_type", "manual_project_subject_type", "project_subject_type"):
        label = normalize_label(row.get(col))
        if label in INCLUDE_LABELS:
            return True
        if label:
            return False

    return False


def run(cmd: List[str], cwd: Optional[Path] = None, check: bool = True, timeout: Optional[int] = None) -> subprocess.CompletedProcess:
    return subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=check,
        timeout=timeout,
    )


def clone_or_update_repo(full_name: str, html_url: str, cache_dir: Path, refresh: bool = False) -> Tuple[Path, str]:
    dest = cache_dir / safe_repo_dir(full_name)

    if refresh and dest.exists():
        shutil.rmtree(dest)

    if dest.exists() and not (dest / ".git").exists():
        shutil.rmtree(dest)

    if not dest.exists():
        clone_url = html_url
        if not clone_url.endswith(".git"):
            clone_url = clone_url.rstrip("/") + ".git"
        print(f"[clone] {full_name}", file=sys.stderr)
        run(["git", "clone", "--depth", "1", clone_url, str(dest)], check=True, timeout=900)
    else:
        print(f"[cache] {full_name}", file=sys.stderr)

    try:
        commit = run(["git", "-C", str(dest), "rev-parse", "HEAD"], check=True).stdout.strip()
    except Exception:
        commit = "HEAD"

    return dest, commit


# ============================================================
# Cache validation (resume support)
# ============================================================

def analyzer_fingerprint(analyzer_path: Path) -> Dict[str, Any]:
    """Build a minimal fingerprint of the current analyzer config. Used to
    detect when cached per-repo outputs are stale because the analyzer file
    or its options have changed since the cache was written.

    Uses mtime + size for change detection (not content hash) — fast and
    sufficient for the common case of "did this file change since the cache
    was written". Uses nanosecond-resolution mtime so two edits within the
    same second don't slip through.
    """
    try:
        st = analyzer_path.stat()
    except OSError:
        return {
            "analyzer_path": str(analyzer_path.resolve()),
            "analyzer_mtime_ns": None,
            "analyzer_size": None,
        }
    return {
        "analyzer_path": str(analyzer_path.resolve()),
        "analyzer_mtime_ns": st.st_mtime_ns,
        "analyzer_size": st.st_size,
    }


def write_per_repo_meta(meta_path: Path, fingerprint: Dict[str, Any], include_low_confidence: bool) -> None:
    """Sidecar metadata for a freshly-written per-repo JSON. Stored separately
    so the analyzer's own JSON output stays untouched."""
    payload = {
        **fingerprint,
        "include_low_confidence": bool(include_low_confidence),
    }
    meta_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def load_per_repo_meta(meta_path: Path) -> Optional[Dict[str, Any]]:
    if not meta_path.exists():
        return None
    try:
        return json.loads(meta_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def cache_validity(
    out_json: Path,
    meta_path: Path,
    current_fingerprint: Dict[str, Any],
    current_include_low_confidence: bool,
) -> Tuple[Optional[Dict[str, Any]], str]:
    """Decide whether to reuse a cached per-repo result.

    Returns (result_dict, reason). result_dict is the parsed JSON if the cache
    is valid and may be reused; None otherwise. `reason` is a short tag used
    for logging and counters: "ok", "missing_json", "corrupt_json",
    "missing_meta", "stale_analyzer", "stale_flag".
    """
    if not out_json.exists():
        return None, "missing_json"

    try:
        raw = out_json.read_text(encoding="utf-8")
    except OSError:
        return None, "corrupt_json"
    if not raw.strip():
        return None, "corrupt_json"
    try:
        result = json.loads(raw)
    except json.JSONDecodeError:
        return None, "corrupt_json"
    if not isinstance(result, dict) or "summary" not in result:
        return None, "corrupt_json"

    meta = load_per_repo_meta(meta_path)
    if meta is None:
        # The JSON is valid but we have no record of what analyzer/flags
        # produced it. Safer to redo than to assume compatibility.
        return None, "missing_meta"

    cached_path = meta.get("analyzer_path")
    cached_mtime_ns = meta.get("analyzer_mtime_ns")
    cached_size = meta.get("analyzer_size")
    if (
        cached_path != current_fingerprint.get("analyzer_path")
        or cached_mtime_ns != current_fingerprint.get("analyzer_mtime_ns")
        or cached_size != current_fingerprint.get("analyzer_size")
    ):
        return None, "stale_analyzer"

    if bool(meta.get("include_low_confidence", False)) != bool(current_include_low_confidence):
        return None, "stale_flag"

    return result, "ok"


# ============================================================
# Output helpers
# ============================================================

def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def flatten_file_record(record: Dict[str, Any]) -> Dict[str, Any]:
    evidence = record.get("evidence") or {}
    return {
        "repo": record.get("repo", ""),
        "repo_url": record.get("repo_url", ""),
        "commit": record.get("commit", ""),
        "file_path": record.get("file_path", ""),
        "file_url": record.get("file_url", ""),
        "language": record.get("language", ""),
        "detected_frameworks": ";".join(record.get("detected_frameworks") or []),
        "file_detected_frameworks": ";".join(record.get("file_detected_frameworks") or []),
        "repo_framework_context": ";".join(record.get("repo_framework_context") or []),
        "local_framework_context": ";".join(record.get("local_framework_context") or []),
        "repo_bdd_context": ";".join(record.get("repo_bdd_context") or []),
        "local_bdd_context": ";".join(record.get("local_bdd_context") or []),
        "file_role": record.get("file_role", ""),
        "confidence": record.get("confidence", ""),
        "classification_reason": record.get("classification_reason", ""),
        "test_case_declaration_count": record.get("test_case_declaration_count", 0),
        "group_or_hook_declaration_count": record.get("group_or_hook_declaration_count", 0),
        "bdd_step_declaration_count": record.get("bdd_step_declaration_count", 0),
        "ui_action_count": record.get("ui_action_count", 0),
        "cypress_custom_command_count": record.get("cypress_custom_command_count", 0),
        "cypress_ui_like_custom_command_count": record.get("cypress_ui_like_custom_command_count", 0),
        "cypress_non_ui_custom_command_count": record.get("cypress_non_ui_custom_command_count", 0),
        "cypress_control_call_count": record.get("cypress_control_call_count", 0),
        "assertion_call_count": record.get("assertion_call_count", 0),
        "evidence_imports": " | ".join(evidence.get("imports") or []),
        "evidence_test_case_declarations": " | ".join(evidence.get("test_case_declarations") or []),
        "evidence_group_or_hook_declarations": " | ".join(evidence.get("group_or_hook_declarations") or []),
        "evidence_bdd_step_declarations": " | ".join(evidence.get("bdd_step_declarations") or []),
        "evidence_bdd_contexts_from_imports": " | ".join(evidence.get("bdd_contexts_from_imports") or []),
        "evidence_playwright_fixture_params": " | ".join(evidence.get("playwright_fixture_params") or []),
        "evidence_ui_actions": " | ".join(evidence.get("ui_actions") or []),
        "evidence_cypress_direct_ui_calls": " | ".join(evidence.get("cypress_direct_ui_calls") or []),
        "evidence_cypress_custom_commands": " | ".join(evidence.get("cypress_custom_commands") or []),
        "evidence_cypress_ui_like_custom_commands": " | ".join(evidence.get("cypress_ui_like_custom_commands") or []),
        "evidence_cypress_non_ui_custom_commands": " | ".join(evidence.get("cypress_non_ui_custom_commands") or []),
        "evidence_cypress_control_calls": " | ".join(evidence.get("cypress_control_calls") or []),
        "evidence_assertion_calls": " | ".join(evidence.get("assertion_calls") or []),
    }


def aggregate_from_result(
    result: Dict[str, Any],
    full_name: str,
    html_url: str,
    commit: str,
    elapsed_seconds: float,
    analysis_source: str,
    all_file_records: List[Dict[str, Any]],
    support_or_setup_records: List[Dict[str, Any]],
    template_file_records: List[Dict[str, Any]],
    low_confidence_records: List[Dict[str, Any]],
    repo_summary_records: List[Dict[str, Any]],
) -> None:
    """Aggregate one repo's analyzer result into the combined lists. Used for
    both freshly-completed analyses and cached results loaded on resume."""
    summary = result.get("summary") or {}
    summary_row = {
        "repo": full_name,
        "repo_url": html_url,
        "commit": commit,
        "detected_ui_test_files": summary.get("detected_ui_test_files", 0),
        "support_or_setup_files": summary.get("support_or_setup_files", 0),
        "template_files": summary.get("template_files", 0),
        "low_confidence_candidates": summary.get("low_confidence_candidates", 0),
        "candidate_test_files": summary.get("candidate_test_files", 0),
        "total_files_scanned": summary.get("total_files_scanned", 0),
        "parse_errors": summary.get("parse_errors", 0),
        "repo_framework_context": " | ".join(summary.get("repo_framework_context") or []),
        "repo_bdd_context": " | ".join(summary.get("repo_bdd_context") or []),
        "framework_distribution": json.dumps(summary.get("framework_distribution") or {}, ensure_ascii=False),
        "confidence_distribution": json.dumps(summary.get("confidence_distribution") or {}, ensure_ascii=False),
        "config_files": " | ".join(summary.get("config_files") or []),
        "config_candidate_dirs": " | ".join(summary.get("config_candidate_dirs") or []),
        "package_script_candidate_dirs": " | ".join(summary.get("package_script_candidate_dirs") or []),
        "elapsed_seconds": round(elapsed_seconds, 2),
        "analysis_source": analysis_source,
    }
    repo_summary_records.append(summary_row)

    for record in result.get("ui_test_files") or []:
        all_file_records.append(flatten_file_record(record))

    for record in result.get("support_or_setup_files") or []:
        support_or_setup_records.append(flatten_file_record(record))

    for record in result.get("template_files") or []:
        template_file_records.append(flatten_file_record(record))

    for record in result.get("low_confidence_candidates") or []:
        low_confidence_records.append(flatten_file_record(record))


# ============================================================
# Main
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(description="Clone repos and discover browser-driven UI test files.")
    parser.add_argument("--repos-csv", type=Path, required=True, help="CSV/XLSX/JSONL with full_name/html_url and optional LLM include labels")
    parser.add_argument("--repo-cache", type=Path, default=Path("repos_cache"))
    parser.add_argument("--output-dir", type=Path, default=Path("ui_file_inventory"))
    parser.add_argument("--node-analyzer", type=Path, required=True, help="Path to analyze_repo_ui_files.cjs")
    parser.add_argument("--only-included", action="store_true", help="Only analyze rows labeled include by the LLM/manual classifier")
    parser.add_argument(
        "--refresh-clones",
        action="store_true",
        help="Delete and re-clone repos. Also invalidates cached per-repo "
             "analyzer outputs so the analyzer is re-run.",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Reuse cached per-repo analyzer outputs when their metadata "
             "(analyzer file + flags) matches the current run. Without this "
             "flag, every repo's analyzer step is re-run and existing "
             "per_repo_outputs/*.json files are overwritten.",
    )
    parser.add_argument("--include-low-confidence", action="store_true", help="Include low-confidence candidates in all_ui_test_files outputs")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--stop-on-error", action="store_true", help="Stop immediately if a repo clone/analyzer step fails")
    args = parser.parse_args()

    if shutil.which("git") is None:
        raise RuntimeError("git not found on PATH")
    if shutil.which("node") is None:
        raise RuntimeError("node not found on PATH")
    if not args.node_analyzer.exists():
        raise FileNotFoundError(args.node_analyzer)

    rows = read_rows(args.repos_csv)

    repos: List[Tuple[str, str]] = []
    seen = set()
    for row in rows:
        if not should_include_repo(row, args.only_included):
            continue
        ident = get_repo_identity(row)
        if not ident:
            continue
        full_name, html_url = ident
        if full_name in seen:
            continue
        seen.add(full_name)
        repos.append((full_name, html_url))

    if args.limit is not None:
        repos = repos[: args.limit]

    args.repo_cache.mkdir(parents=True, exist_ok=True)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    per_repo_dir = args.output_dir / "per_repo_outputs"
    per_repo_dir.mkdir(parents=True, exist_ok=True)

    # Capture the analyzer fingerprint once at startup. Used both to write
    # fresh meta files and to validate cached ones.
    current_fingerprint = analyzer_fingerprint(args.node_analyzer)

    print(
        f"[config] resume={args.resume} refresh_clones={args.refresh_clones} "
        f"include_low_confidence={args.include_low_confidence}",
        file=sys.stderr,
    )

    all_file_records: List[Dict[str, Any]] = []
    support_or_setup_records: List[Dict[str, Any]] = []
    template_file_records: List[Dict[str, Any]] = []
    low_confidence_records: List[Dict[str, Any]] = []
    repo_summary_records: List[Dict[str, Any]] = []
    error_records: List[Dict[str, Any]] = []

    resumed_count = 0
    fresh_count = 0
    stale_cache_reasons: Dict[str, int] = {}

    for idx, (full_name, html_url) in enumerate(repos, start=1):
        out_json = per_repo_dir / f"{safe_repo_dir(full_name)}.json"
        meta_json = per_repo_dir / f"{safe_repo_dir(full_name)}.meta.json"

        # --refresh-clones invalidates both clone and per-repo cache, regardless
        # of --resume. The deletion happens up front so the resume check below
        # finds an empty slate.
        if args.refresh_clones:
            for stale in (out_json, meta_json):
                if stale.exists():
                    try:
                        stale.unlink()
                    except OSError:
                        pass

        # Only consult the cache when --resume is explicitly set. Default
        # behavior is to re-run the analyzer and overwrite per-repo outputs.
        if args.resume and not args.refresh_clones:
            cached, reason = cache_validity(
                out_json,
                meta_json,
                current_fingerprint,
                args.include_low_confidence,
            )
            if cached is not None:
                resumed_count += 1
                print(f"[{idx}/{len(repos)}] {full_name} [resume from cache]", file=sys.stderr)
                cached_summary = cached.get("summary") or {}
                cached_commit = cached.get("commit") or cached_summary.get("commit") or "HEAD"
                cached_elapsed = cached_summary.get("elapsed_seconds", 0.0)
                aggregate_from_result(
                    cached,
                    full_name=full_name,
                    html_url=html_url,
                    commit=cached_commit,
                    elapsed_seconds=float(cached_elapsed or 0.0),
                    analysis_source="cached",
                    all_file_records=all_file_records,
                    support_or_setup_records=support_or_setup_records,
                    template_file_records=template_file_records,
                    low_confidence_records=low_confidence_records,
                    repo_summary_records=repo_summary_records,
                )
                continue
            # Cache existed in some form but isn't valid for this run.
            # Log the reason and fall through to the fresh analysis path.
            if reason != "missing_json":
                stale_cache_reasons[reason] = stale_cache_reasons.get(reason, 0) + 1
                print(
                    f"[{idx}/{len(repos)}] {full_name} [stale cache: {reason}, re-running]",
                    file=sys.stderr,
                )

        # Fresh clone + analyze path.
        # Delete any pre-existing per-repo JSON + meta first. Without this,
        # if the analyzer fails after this point, the old files survive on
        # disk and could be picked up by a later --resume run, producing
        # silently stale results paired with the new analyzer config.
        # Note: --refresh-clones already deletes these earlier; this is a
        # safety net for the default no-cache path.
        for stale in (out_json, meta_json):
            if stale.exists():
                try:
                    stale.unlink()
                except OSError:
                    pass

        print(f"[{idx}/{len(repos)}] {full_name}", file=sys.stderr)
        started = time.time()

        try:
            local_path, commit = clone_or_update_repo(full_name, html_url, args.repo_cache, refresh=args.refresh_clones)

            cmd = [
                "node",
                str(args.node_analyzer),
                "--repo-path", str(local_path),
                "--repo", full_name,
                "--repo-url", html_url,
                "--commit", commit,
                "--output", str(out_json),
            ]
            if args.include_low_confidence:
                cmd.append("--include-low-confidence")

            run(cmd, check=True, timeout=900)

            # Write the sidecar metadata so a future --resume run can validate
            # this output against its analyzer + flags.
            write_per_repo_meta(meta_json, current_fingerprint, args.include_low_confidence)

            result = json.loads(out_json.read_text(encoding="utf-8"))
            aggregate_from_result(
                result,
                full_name=full_name,
                html_url=html_url,
                commit=commit,
                elapsed_seconds=time.time() - started,
                analysis_source="fresh",
                all_file_records=all_file_records,
                support_or_setup_records=support_or_setup_records,
                template_file_records=template_file_records,
                low_confidence_records=low_confidence_records,
                repo_summary_records=repo_summary_records,
            )
            fresh_count += 1

        except subprocess.CalledProcessError as exc:
            error_records.append({
                "repo": full_name,
                "repo_url": html_url,
                "error": repr(exc),
                "stdout": (exc.stdout or "")[-2000:],
                "stderr": (exc.stderr or "")[-2000:],
                "elapsed_seconds": round(time.time() - started, 2),
            })
            print(f"[error] {full_name}: {exc}", file=sys.stderr)
            if args.stop_on_error:
                raise
        except Exception as exc:
            error_records.append({
                "repo": full_name,
                "repo_url": html_url,
                "error": repr(exc),
                "elapsed_seconds": round(time.time() - started, 2),
            })
            print(f"[error] {full_name}: {exc}", file=sys.stderr)
            if args.stop_on_error:
                raise

    file_fieldnames = [
        "repo", "repo_url", "commit", "file_path", "file_url", "language",
        "detected_frameworks", "file_detected_frameworks", "repo_framework_context", "local_framework_context", "repo_bdd_context", "local_bdd_context",
        "file_role", "confidence", "classification_reason", "test_case_declaration_count",
        "group_or_hook_declaration_count", "bdd_step_declaration_count", "ui_action_count", "cypress_custom_command_count", "cypress_ui_like_custom_command_count", "cypress_non_ui_custom_command_count", "cypress_control_call_count", "assertion_call_count",
        "evidence_imports", "evidence_test_case_declarations",
        "evidence_group_or_hook_declarations", "evidence_bdd_step_declarations", "evidence_bdd_contexts_from_imports", "evidence_playwright_fixture_params",
        "evidence_ui_actions", "evidence_cypress_direct_ui_calls", "evidence_cypress_custom_commands", "evidence_cypress_ui_like_custom_commands", "evidence_cypress_non_ui_custom_commands", "evidence_cypress_control_calls", "evidence_assertion_calls",
    ]
    summary_fieldnames = [
        "repo", "repo_url", "commit", "detected_ui_test_files", "support_or_setup_files", "template_files", "low_confidence_candidates",
        "candidate_test_files", "total_files_scanned", "parse_errors", "repo_framework_context", "repo_bdd_context",
        "framework_distribution", "confidence_distribution", "config_files",
        "config_candidate_dirs", "package_script_candidate_dirs", "elapsed_seconds",
        "analysis_source",
    ]

    write_csv(args.output_dir / "all_ui_test_files.csv", all_file_records, file_fieldnames)
    write_jsonl(args.output_dir / "all_ui_test_files.jsonl", all_file_records)
    write_csv(args.output_dir / "support_or_setup_files.csv", support_or_setup_records, file_fieldnames)
    write_jsonl(args.output_dir / "support_or_setup_files.jsonl", support_or_setup_records)
    write_csv(args.output_dir / "template_files.csv", template_file_records, file_fieldnames)
    write_jsonl(args.output_dir / "template_files.jsonl", template_file_records)
    write_csv(args.output_dir / "low_confidence_candidates.csv", low_confidence_records, file_fieldnames)
    write_jsonl(args.output_dir / "low_confidence_candidates.jsonl", low_confidence_records)
    write_csv(args.output_dir / "repo_ui_file_summary.csv", repo_summary_records, summary_fieldnames)
    write_jsonl(args.output_dir / "repo_ui_file_summary.jsonl", repo_summary_records)
    write_jsonl(args.output_dir / "errors.jsonl", error_records)

    overall_summary = {
        "input_rows": len(rows),
        "repos_selected_for_analysis": len(repos),
        "repos_processed_successfully": len(repo_summary_records),
        "repos_resumed_from_cache": resumed_count,
        "repos_newly_processed": fresh_count,
        "repos_skipped_stale_cache": stale_cache_reasons,
        "repos_with_errors": len(error_records),
        "detected_ui_test_files": len(all_file_records),
        "support_or_setup_files": len(support_or_setup_records),
        "template_files": len(template_file_records),
        "low_confidence_candidates": len(low_confidence_records),
        "output_dir": str(args.output_dir),
        "config": {
            "resume": args.resume,
            "refresh_clones": args.refresh_clones,
            "include_low_confidence": args.include_low_confidence,
            "analyzer_fingerprint": current_fingerprint,
        },
    }
    (args.output_dir / "overall_summary.json").write_text(
        json.dumps(overall_summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(json.dumps(overall_summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
